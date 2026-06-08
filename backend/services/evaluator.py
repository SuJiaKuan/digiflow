"""GT 評估模組：載入 xlsx、比對辨識結果、計算 FAR (Recall / Precision / F1)。

GT 檔案存放於 test_data/A05/print/，命名規則：
  image:  A05-P-N-260116-001.jpg   (第三段為 N)
  GT:     A05-P-B-260116-001.xlsx  (第三段為 B)

General 分頁：key-value 表，比對 維修廠 / 估價日期 / 牌照號碼 / VIN / 總計。
Data 分頁   ：第一列為表頭，往下逐列比對。
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

_GT_DIR = Path(__file__).parent.parent.parent / "test_data" / "A05" / "print"

# 辨識結果固定 meta 欄位，不視為 Data 欄位
_META_COLS = {"來源檔名", "維修廠", "估價日期", "牌照號碼"}

# General 分頁中跳過的 key（純 meta）
_GENERAL_SKIP = {"來源檔名", "Key"}


# ── 值正規化 ───────────────────────────────────────────────

def _norm_col(s: Any) -> str:
    return str(s).strip() if s is not None else ""


def _norm_gt_value(v: Any) -> str:
    """GT xlsx cell → 可比較字串。"""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y/%m/%d")
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return str(round(v, 6))
    return str(v).strip()


def _norm_pred_value(v: Any) -> str:
    """辨識結果 JSON 值 → 可比較字串；嘗試把 '543.0' 轉成 '543'。"""
    if v is None:
        return ""
    s = str(v).strip()
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except (ValueError, OverflowError):
        pass
    return s


# ── GT 路徑查找 ────────────────────────────────────────────

def find_gt_path(filename: str) -> Path | None:
    """根據上傳檔名找對應 GT xlsx。
    規則：把第三個 dash 分段由 N 換成 B，副檔名換成 .xlsx。
    找不到時 fallback：掃描目錄比對 General['來源檔名']。
    """
    stem = Path(filename).stem
    parts = stem.split("-")
    if len(parts) >= 5:
        candidate = parts.copy()
        candidate[2] = "B"
        gt_path = _GT_DIR / f"{'-'.join(candidate)}.xlsx"
        if gt_path.exists():
            return gt_path

    # fallback：逐一讀 General 分頁確認 來源檔名
    if _GT_DIR.exists():
        for f in sorted(_GT_DIR.glob("*.xlsx")):
            try:
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                if "General" in wb.sheetnames:
                    for row in wb["General"].iter_rows(values_only=True):
                        if row and len(row) >= 2 and row[0] == "來源檔名":
                            if str(row[1]).strip() == filename:
                                wb.close()
                                return f
                wb.close()
            except Exception:
                pass
    return None


# ── GT 載入 ────────────────────────────────────────────────

def _load_gt(gt_path: Path) -> dict:
    """回傳 {'general': dict[str,str], 'data': list[dict[str,str]]}。"""
    wb = openpyxl.load_workbook(gt_path, read_only=True, data_only=True)

    # General 分頁
    general: dict[str, str] = {}
    if "General" in wb.sheetnames:
        for row in wb["General"].iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip()
            if key in _GENERAL_SKIP:
                continue
            val = row[1] if len(row) >= 2 else None
            general[key] = _norm_gt_value(val)

    # Data 分頁：取第一個非 General 的分頁
    data_ws = None
    for name in wb.sheetnames:
        if name != "General":
            data_ws = wb[name]
            break

    data: list[dict[str, str]] = []
    if data_ws is not None:
        rows = list(data_ws.iter_rows(values_only=True))
        if rows:
            # 第一列為表頭，過濾掉 None 的欄位名
            header_raw = rows[0]
            header = [_norm_col(c) for c in header_raw if c is not None]
            n_cols = len(header)
            for row in rows[1:]:
                if not any(v is not None for v in row):
                    continue
                row_dict: dict[str, str] = {}
                for i, col in enumerate(header):
                    if not col:
                        continue
                    val = row[i] if i < len(row) else None
                    row_dict[col] = _norm_gt_value(val)
                data.append(row_dict)

    wb.close()
    return {"general": general, "data": data}


# ── 辨識結果拆解 ───────────────────────────────────────────

def _extract_prediction(recognition_rows: list[dict]) -> dict:
    """把辨識結果 flat rows → {'general': dict, 'data': list[dict]}。"""
    if not recognition_rows:
        return {"general": {}, "data": []}

    first = recognition_rows[0]
    general: dict[str, str] = {}
    for key in ("維修廠", "估價日期", "牌照號碼"):
        general[key] = _norm_pred_value(first.get(key))

    data: list[dict[str, str]] = []
    for row in recognition_rows:
        row_dict = {
            _norm_col(k): _norm_pred_value(v)
            for k, v in row.items()
            if k not in _META_COLS
        }
        data.append(row_dict)

    return {"general": general, "data": data}


# ── 比對邏輯 ───────────────────────────────────────────────

def _compare_general(gt: dict[str, str], pred: dict[str, str]) -> dict:
    wrong: list[dict] = []
    missing: list[dict] = []
    extra: list[dict] = []
    C = 0

    # 用 normalized key 做比對
    gt_norm = {_norm_col(k): (k, v) for k, v in gt.items()}
    pred_norm = {_norm_col(k): (k, v) for k, v in pred.items()}
    all_norm = sorted(set(gt_norm) | set(pred_norm))

    for nk in all_norm:
        in_gt = nk in gt_norm
        in_pred = nk in pred_norm
        gt_key, gt_val = gt_norm[nk] if in_gt else (nk, "")
        pred_key, pred_val = pred_norm[nk] if in_pred else (nk, "")

        if in_gt and in_pred:
            if gt_val == pred_val:
                C += 1
            else:
                wrong.append({"key": gt_key, "gt": gt_val, "pred": pred_val})
        elif in_gt:
            missing.append({"key": gt_key, "gt": gt_val})
        else:
            extra.append({"key": pred_key, "pred": pred_val})

    return {
        "wrong": wrong,
        "missing": missing,
        "extra": extra,
        "C": C,
        "G": len(gt),
        "P": len(pred),
    }


def _compare_data(gt_data: list[dict], pred_data: list[dict]) -> dict:
    wrong: list[dict] = []
    missing: list[dict] = []
    extra: list[dict] = []
    C = G = P = 0

    # 建立全域正規化欄位名 → 原始名稱的映射（以 GT 為主，pred 為輔）
    gt_col_map: dict[str, str] = {}
    for row in gt_data:
        for c in row:
            gt_col_map.setdefault(_norm_col(c), c)
    pred_col_map: dict[str, str] = {}
    for row in pred_data:
        for c in row:
            pred_col_map.setdefault(_norm_col(c), c)

    max_rows = max(len(gt_data), len(pred_data))

    for idx in range(max_rows):
        row_num = idx + 1  # 1-based，顯示用
        gt_row = gt_data[idx] if idx < len(gt_data) else {}
        pred_row = pred_data[idx] if idx < len(pred_data) else {}

        gt_row_norm = {_norm_col(k): v for k, v in gt_row.items()}
        pred_row_norm = {_norm_col(k): v for k, v in pred_row.items()}

        all_norm_cols = sorted(set(gt_row_norm) | set(pred_row_norm))

        for nc in all_norm_cols:
            in_gt = nc in gt_row_norm
            in_pred = nc in pred_row_norm
            gt_val = gt_row_norm.get(nc, "")
            pred_val = pred_row_norm.get(nc, "")

            if in_gt:
                G += 1
            if in_pred:
                P += 1

            if in_gt and in_pred:
                if gt_val == pred_val:
                    C += 1
                else:
                    wrong.append({
                        "row": row_num,
                        "col": gt_col_map.get(nc, nc),
                        "gt": gt_val,
                        "pred": pred_val,
                    })
            elif in_gt:
                missing.append({
                    "row": row_num,
                    "col": gt_col_map.get(nc, nc),
                    "gt": gt_val,
                })
            else:
                extra.append({
                    "row": row_num,
                    "col": pred_col_map.get(nc, nc),
                    "pred": pred_val,
                })

    return {"wrong": wrong, "missing": missing, "extra": extra, "C": C, "G": G, "P": P}


# ── 主入口 ─────────────────────────────────────────────────

def evaluate(recognition_rows: list[dict], filename: str) -> dict | None:
    """找 GT、比對、回傳評估結果；找不到 GT 則回傳 None。"""
    gt_path = find_gt_path(filename)
    if gt_path is None:
        return None

    try:
        gt = _load_gt(gt_path)
    except Exception as exc:
        return {"error": f"GT 載入失敗：{exc}"}

    try:
        pred = _extract_prediction(recognition_rows)
        gen = _compare_general(gt["general"], pred["general"])
        data = _compare_data(gt["data"], pred["data"])
    except Exception as exc:
        return {"error": f"比對失敗：{exc}"}

    C = gen["C"] + data["C"]
    G = gen["G"] + data["G"]
    P = gen["P"] + data["P"]
    recall = round(C / G * 100, 1) if G > 0 else 0.0
    precision = round(C / P * 100, 1) if P > 0 else 0.0
    f1 = (
        round(2 * precision * recall / (precision + recall), 1)
        if (precision + recall) > 0
        else 0.0
    )

    return {
        "scores": {
            "C": C, "G": G, "P": P,
            "recall": recall,
            "precision": precision,
            "f1": f1,
        },
        "general": {
            "wrong": gen["wrong"],
            "missing": gen["missing"],
            "extra": gen["extra"],
        },
        "data": {
            "wrong": data["wrong"],
            "missing": data["missing"],
            "extra": data["extra"],
        },
    }
